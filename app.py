import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import io
import re
from datetime import datetime, date
import warnings
from typing import List, Dict, Any, Optional, Tuple
import json

# Suppress warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Page configuration
st.set_page_config(
    page_title="Robust Bank Search",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .success-box {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        color: #155724;
        margin: 1rem 0;
    }
    .warning-box {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #fff3cd;
        border: 1px solid #ffeeba;
        color: #856404;
        margin: 1rem 0;
    }
    .info-box {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #d1ecf1;
        border: 1px solid #bee5eb;
        color: #0c5460;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

# Debug logging
if 'search_log' not in st.session_state:
    st.session_state.search_log = []

def add_log(message: str):
    """Add message to search log"""
    timestamp = datetime.now().strftime('%H:%M:%S')
    st.session_state.search_log.append(f"[{timestamp}] {message}")

def clean_for_excel(value) -> str:
    """Clean value for Excel compatibility"""
    if pd.isna(value) or value is None:
        return ""
    
    # Convert to string and clean
    str_value = str(value)
    
    # Remove or replace problematic characters
    str_value = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F-\x9F]', '', str_value)  # Remove control characters
    str_value = str_value.replace('\n', ' ').replace('\r', ' ')  # Replace line breaks
    str_value = str_value.strip()
    
    # Limit length to prevent Excel issues
    if len(str_value) > 32000:  # Excel cell limit is ~32,767 characters
        str_value = str_value[:32000] + "..."
    
    return str_value

def clean_numeric_value(value) -> Optional[float]:
    """Extract numeric value from string"""
    if isinstance(value, (int, float)):
        return float(value) if not pd.isna(value) else None
    
    if isinstance(value, str) and value.strip():
        cleaned = re.sub(r'[^\d.-]', '', value.replace(',', ''))
        try:
            return float(cleaned) if cleaned else None
        except ValueError:
            return None
    return None

def safe_date_to_string(date_value) -> str:
    """Convert date to Excel-safe string"""
    if pd.isna(date_value) or date_value is None:
        return ""
    
    if isinstance(date_value, (date, datetime)):
        return date_value.strftime('%Y-%m-%d')
    
    return clean_for_excel(date_value)

def parse_date(value) -> Optional[date]:
    """Parse date from various formats"""
    if pd.isna(value) or value == '':
        return None
    
    if isinstance(value, (date, datetime)):
        return value.date() if isinstance(value, datetime) else value
    
    if isinstance(value, str):
        # Try common date formats
        date_formats = ['%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d']
        for fmt in date_formats:
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    
    return None

def sanitize_sheet_name(name: str) -> str:
    """Sanitize sheet name for Excel compatibility"""
    if not name:
        return "Sheet1"
    
    # Remove invalid characters for Excel sheet names
    sanitized = re.sub(r'[\\/*?[\]:]+', '_', str(name))
    sanitized = clean_for_excel(sanitized)
    
    # Excel sheet names must be 1-31 characters
    if len(sanitized) > 31:
        sanitized = sanitized[:31]
    
    # Cannot be empty
    if not sanitized:
        sanitized = "Sheet1"
    
    return sanitized

def read_bank_amount_data(file_content: bytes, filename: str) -> List[Dict[str, Any]]:
    """Read bank, amount, and date data from input file"""
    try:
        add_log(f"Reading bank-amount data from {filename}")
        
        if filename.endswith('.xlsx'):
            df = pd.read_excel(io.BytesIO(file_content), engine='openpyxl')
        else:
            df = pd.read_excel(io.BytesIO(file_content), engine='xlrd')
        
        add_log(f"Input file shape: {df.shape}")
        
        bank_data = []
        
        # Assuming columns are A=Bank, B=Amount, C=Date
        for idx, row in df.iterrows():
            try:
                bank_name = clean_for_excel(row.iloc[0]).strip().upper() if not pd.isna(row.iloc[0]) else ""
                amount = clean_numeric_value(row.iloc[1]) if len(row) > 1 else None
                search_date = parse_date(row.iloc[2]) if len(row) > 2 else None
                
                if bank_name and amount and amount > 0:
                    bank_data.append({
                        'Bank': bank_name,
                        'Amount': amount,
                        'Date': search_date,
                        'Row': idx + 1
                    })
                    add_log(f"Row {idx+1}: {bank_name} - {amount:,.2f} - {search_date}")
                else:
                    add_log(f"Row {idx+1}: Skipped (invalid data)")
                    
            except Exception as e:
                add_log(f"Error processing row {idx+1}: {e}")
                continue
        
        add_log(f"Total valid bank-amount entries: {len(bank_data)}")
        return bank_data
        
    except Exception as e:
        add_log(f"Error reading input file: {e}")
        st.error(f"Error reading input file: {e}")
        return []

def read_excel_file_with_dates(file_content: bytes, filename: str) -> Optional[pd.DataFrame]:
    """Read Excel file and parse dates with robust error handling"""
    try:
        add_log(f"Reading source file: {filename}")
        
        if filename.endswith('.xlsx'):
            wb = load_workbook(io.BytesIO(file_content), read_only=True)
            all_data = []
            
            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    
                    # Convert to DataFrame with size limits
                    data = []
                    max_rows = 10000  # Limit rows to prevent memory issues
                    row_count = 0
                    
                    for row in ws.iter_rows(values_only=True):
                        if row_count >= max_rows:
                            add_log(f"Sheet '{sheet_name}' truncated at {max_rows} rows")
                            break
                        data.append(row)
                        row_count += 1
                    
                    if not data:
                        continue
                        
                    df = pd.DataFrame(data)
                    
                    # Clean column names
                    if len(df) > 0:
                        first_row = df.iloc[0]
                        if any(isinstance(val, str) and len(str(val)) > 0 for val in first_row):
                            df.columns = [clean_for_excel(col) if not pd.isna(col) and col != '' else f"Column_{i}" 
                                        for i, col in enumerate(first_row)]
                            df = df.drop(index=0).reset_index(drop=True)
                        else:
                            df.columns = [f"Column_{i}" for i in range(len(df.columns))]
                    
                    # Clean all data
                    for col in df.columns:
                        df[col] = df[col].apply(lambda x: clean_for_excel(x) if not isinstance(x, (int, float)) else x)
                    
                    # Add metadata
                    df['Source_Sheet'] = clean_for_excel(sheet_name)
                    df['Source_File'] = clean_for_excel(filename)
                    
                    # Try to identify and parse date columns
                    for col in df.columns:
                        if any(keyword in str(col).lower() for keyword in ['date', 'time', 'تاريخ']):
                            add_log(f"Found potential date column: {col}")
                            df[f'{col}_Parsed'] = df[col].apply(parse_date)
                    
                    all_data.append(df)
                    add_log(f"Processed sheet '{sheet_name}' with {len(df)} rows")
                    
                except Exception as e:
                    add_log(f"Error reading sheet '{sheet_name}': {e}")
                    continue
            
            wb.close()
            
            if all_data:
                result = pd.concat(all_data, ignore_index=True)
                add_log(f"Combined data from {filename}: {result.shape}")
                return result
            else:
                return None
                
        elif filename.endswith('.xls'):
            try:
                all_sheets = pd.read_excel(io.BytesIO(file_content), sheet_name=None, engine='xlrd')
                all_data = []
                
                for sheet_name, df in all_sheets.items():
                    # Limit size
                    if len(df) > 10000:
                        df = df.head(10000)
                        add_log(f"Sheet '{sheet_name}' truncated to 10000 rows")
                    
                    # Clean data
                    for col in df.columns:
                        if df[col].dtype == 'object':
                            df[col] = df[col].apply(lambda x: clean_for_excel(x))
                    
                    df['Source_Sheet'] = clean_for_excel(sheet_name)
                    df['Source_File'] = clean_for_excel(filename)
                    
                    # Try to parse date columns
                    for col in df.columns:
                        if any(keyword in str(col).lower() for keyword in ['date', 'time', 'تاريخ']):
                            df[f'{col}_Parsed'] = df[col].apply(parse_date)
                    
                    all_data.append(df)
                    add_log(f"Processed sheet '{sheet_name}' with {len(df)} rows")
                
                if all_data:
                    result = pd.concat(all_data, ignore_index=True)
                    return result
                else:
                    return None
                    
            except Exception as e:
                add_log(f"Error reading .xls file: {e}")
                return None
                
    except Exception as e:
        add_log(f"Error reading file {filename}: {e}")
        return None

def check_bank_match(filename: str, bank_name: str) -> bool:
    """Check if filename/path contains the bank name"""
    filename_upper = filename.upper()
    bank_upper = bank_name.upper()
    
    # Direct match
    if bank_upper in filename_upper:
        return True
    
    # Common bank name variations
    bank_variations = {
        'SNB': ['SNB', 'SAUDI', 'NATIONAL'],
        'RIB': ['RIB', 'RIYADH', 'RIYAD'],
        'SABB': ['SABB', 'SAUDI', 'BRITISH'],
        'INM': ['INM', 'INMA', 'ALINMA'],
        'ARB': ['ARB', 'ARAB', 'NATIONAL'],
        'NCB': ['NCB', 'AHLI', 'NATIONAL'],
        'BSF': ['BSF', 'BANQUE', 'SAUDI', 'FRANSI']
    }
    
    if bank_upper in bank_variations:
        for variation in bank_variations[bank_upper]:
            if variation in filename_upper:
                return True
    
    return False

def find_bank_amount_matches(data: pd.DataFrame, bank_entry: Dict[str, Any], tolerance: float) -> List[Dict[str, Any]]:
    """Find matches for specific bank amount with date filtering"""
    bank_name = bank_entry['Bank']
    target_amount = bank_entry['Amount']
    search_date = bank_entry['Date']
    
    add_log(f"Searching for {bank_name} - {target_amount:,.2f} from {search_date}")
    
    matches = []
    
    for idx, row in data.iterrows():
        # Check amount match first
        amount_matched = False
        matched_columns = []
        matched_values = []
        
        for col_name, value in row.items():
            if col_name in ['Source_Sheet', 'Source_File']:
                continue
                
            numeric_value = clean_numeric_value(value)
            
            if numeric_value is not None:
                if abs(numeric_value - target_amount) <= tolerance:
                    amount_matched = True
                    matched_columns.append(clean_for_excel(col_name))
                    matched_values.append(f"{numeric_value:,.2f}")
                    break
        
        if amount_matched:
            # Check date condition if date is specified
            date_condition_met = True
            transaction_date = None
            
            if search_date:
                # Look for date columns in the row
                date_columns = [col for col in data.columns if 'date' in col.lower() or col.endswith('_Parsed')]
                
                for date_col in date_columns:
                    row_date = parse_date(row[date_col])
                    if row_date and row_date >= search_date:
                        transaction_date = row_date
                        break
                else:
                    # If no suitable date found, check if we should still include
                    if date_columns:  # Has date columns but none meet criteria
                        date_condition_met = False
                        add_log(f"Row {idx+1}: Amount matched but date condition not met")
            
            if date_condition_met:
                # Create clean match record
                clean_match = {
                    'Bank_Searched': clean_for_excel(bank_name),
                    'Search_Amount': target_amount,
                    'Search_Date': safe_date_to_string(search_date),
                    'Transaction_Date': safe_date_to_string(transaction_date),
                    'Matched_Columns': ', '.join(matched_columns),
                    'Matched_Values': ', '.join(matched_values),
                    'Tolerance_Used': tolerance,
                    'Search_Timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'Row_Number': idx + 1,
                    'Source_File': clean_for_excel(row.get('Source_File', '')),
                    'Source_Sheet': clean_for_excel(row.get('Source_Sheet', ''))
                }
                
                # Add other row data (cleaned)
                for key, value in row.items():
                    if key not in clean_match and not key.endswith('_Parsed'):
                        clean_key = clean_for_excel(key)
                        clean_value = clean_for_excel(value) if not isinstance(value, (int, float)) else value
                        clean_match[clean_key] = clean_value
                
                matches.append(clean_match)
                add_log(f"MATCH: {bank_name} - {target_amount:,.2f} found in row {idx+1}")
    
    return matches

def process_bank_specific_search(bank_data: List[Dict[str, Any]], source_files: List, tolerance: float) -> Tuple[Dict[str, List[Dict[str, Any]]], Dict[str, Dict[str, Any]]]:
    """Process bank-specific search with robust error handling"""
    add_log(f"Starting bank-specific search for {len(bank_data)} entries")
    
    all_matches = {}
    search_summary = {}
    
    # Initialize summary
    for entry in bank_data:
        key = f"{entry['Bank']}_{entry['Amount']:,.2f}"
        search_summary[key] = {
            'Bank': entry['Bank'],
            'Amount': entry['Amount'],
            'Date': safe_date_to_string(entry['Date']),
            'Matches_Found': 0,
            'Files_Searched': [],
            'Status': 'Not Found'
        }
    
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    # Group bank data by bank for efficient processing
    banks_to_search = {}
    for entry in bank_data:
        bank = entry['Bank']
        if bank not in banks_to_search:
            banks_to_search[bank] = []
        banks_to_search[bank].append(entry)
    
    add_log(f"Banks to search: {list(banks_to_search.keys())}")
    
    total_operations = len(source_files) * len(banks_to_search)
    current_operation = 0
    
    for uploaded_file in source_files:
        add_log(f"Processing file: {uploaded_file.name}")
        
        try:
            file_content = uploaded_file.read()
            data = read_excel_file_with_dates(file_content, uploaded_file.name)
            
            if data is not None and not data.empty:
                add_log(f"File loaded: {len(data)} rows")
                
                # Check which banks this file should be searched for
                for bank_name, bank_entries in banks_to_search.items():
                    current_operation += 1
                    status_text.text(f"Searching {bank_name} in {uploaded_file.name}...")
                    
                    if check_bank_match(uploaded_file.name, bank_name):
                        add_log(f"File {uploaded_file.name} matches bank {bank_name} - searching...")
                        
                        # Search for all amounts for this bank in this file
                        for bank_entry in bank_entries:
                            key = f"{bank_entry['Bank']}_{bank_entry['Amount']:,.2f}"
                            search_summary[key]['Files_Searched'].append(uploaded_file.name)
                            
                            matches = find_bank_amount_matches(data, bank_entry, tolerance)
                            
                            if matches:
                                file_key = sanitize_sheet_name(f"{bank_name}_{uploaded_file.name.split('.')[0]}")
                                
                                if file_key not in all_matches:
                                    all_matches[file_key] = []
                                all_matches[file_key].extend(matches)
                                
                                search_summary[key]['Matches_Found'] += len(matches)
                                search_summary[key]['Status'] = 'Found'
                                
                                add_log(f"Added {len(matches)} matches for {key}")
                    else:
                        add_log(f"File {uploaded_file.name} does not match bank {bank_name} - skipping")
                    
                    progress_bar.progress(current_operation / total_operations)
            else:
                add_log(f"Could not load file: {uploaded_file.name}")
                
        except Exception as e:
            add_log(f"Error processing {uploaded_file.name}: {e}")
            st.error(f"Error processing {uploaded_file.name}: {e}")
    
    status_text.text("Search complete!")
    progress_bar.empty()
    status_text.empty()
    
    add_log(f"Search completed. Found matches in {len(all_matches)} file-bank combinations")
    
    return all_matches, search_summary

def create_robust_excel_report(matches_dict: Dict[str, List[Dict[str, Any]]], search_summary: Dict[str, Dict[str, Any]], bank_data: List[Dict[str, Any]]) -> bytes:
    """Create Excel report with only result sheets - no summary/log sheets"""
    output = io.BytesIO()
    
    try:
        with pd.ExcelWriter(output, engine='openpyxl', options={'remove_timezone': True}) as writer:
            
            # Only create individual result sheets - no summary sheets
            sheet_count = 0
            max_sheets = 50  # Limit number of sheets
            
            for file_key, matches in matches_dict.items():
                if sheet_count >= max_sheets:
                    add_log(f"Reached maximum sheet limit ({max_sheets})")
                    break
                
                if matches:
                    try:
                        # Create clean dataframe
                        clean_matches = []
                        for match in matches[:1000]:  # Limit rows per sheet
                            clean_match = {}
                            for key, value in match.items():
                                clean_key = clean_for_excel(key)[:30]  # Limit column name length
                                if isinstance(value, (int, float)):
                                    clean_match[clean_key] = value
                                else:
                                    clean_match[clean_key] = clean_for_excel(value)
                            clean_matches.append(clean_match)
                        
                        if clean_matches:
                            df = pd.DataFrame(clean_matches)
                            
                            # Ensure column order
                            priority_cols = ['Bank_Searched', 'Search_Amount', 'Search_Date', 'Transaction_Date',
                                           'Matched_Columns', 'Matched_Values', 'Source_File', 'Source_Sheet', 'Row_Number']
                            
                            available_priority = [col for col in priority_cols if col in df.columns]
                            other_cols = [col for col in df.columns if col not in priority_cols]
                            
                            # Limit total columns
                            all_cols = available_priority + other_cols[:20]  # Max 20 additional columns
                            df = df[all_cols]
                            
                            # Clean sheet name
                            sheet_name = sanitize_sheet_name(file_key)
                            
                            # Ensure unique sheet name
                            original_sheet_name = sheet_name
                            counter = 1
                            while sheet_name in [sheet.title for sheet in writer.book.worksheets] if hasattr(writer, 'book') else False:
                                sheet_name = f"{original_sheet_name[:25]}_{counter}"
                                counter += 1
                            
                            df.to_excel(writer, sheet_name=sheet_name, index=False)
                            sheet_count += 1
                            
                    except Exception as e:
                        add_log(f"Error creating sheet for {file_key}: {e}")
                        continue
        
        add_log("Excel file created successfully")
        
    except Exception as e:
        add_log(f"Error creating Excel file: {e}")
        # Create empty output if error occurs
        st.error(f"Failed to create Excel file: {e}")
        return b""
    
    output.seek(0)
    return output.getvalue()

def main():
    """Main Streamlit app"""
    
    # Header
    st.markdown('<h1 class="main-header">🏦 Robust Bank Collection Search</h1>', unsafe_allow_html=True)
    
    # Introduction
    st.markdown("""
    <div class="info-box">
    🛡️ <strong>Excel-Safe Bank Search Tool</strong><br>
    Enhanced version with robust Excel export handling to prevent file corruption issues.
    Upload bank-amount-date data and search through bank statement files.
    </div>
    """, unsafe_allow_html=True)
    
    # Sidebar
    with st.sidebar:
        st.header("🔧 Search Configuration")
        
        tolerance = st.number_input(
            "Search tolerance (±):",
            min_value=0.00,
            value=0.01,
            step=0.01,
            format="%.2f",
            help="Allow for small differences in the amount"
        )
        
        st.markdown("---")
        st.markdown("### 📋 Input File Format")
        st.markdown("""
        **Column A**: Bank name (SNB, RIB, SABB, etc.)  
        **Column B**: Amount to search for  
        **Column C**: Date (search on/after this date)
        """)
        
        st.markdown("---")
        if st.button("Clear Search Log"):
            st.session_state.search_log = []
            st.success("Log cleared!")
        
        show_log = st.checkbox("Show Search Log")
        if show_log and st.session_state.search_log:
            st.text_area("Recent Log", "\n".join(st.session_state.search_log[-10:]), height=200)
    
    # File upload sections
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📥 Input File (Bank-Amount-Date)")
        
        input_file = st.file_uploader(
            "Upload Excel file with Bank, Amount, Date columns",
            type=['xlsx', 'xls'],
            help="Format: Column A = Bank, Column B = Amount, Column C = Date"
        )
        
        if input_file:
            st.success(f"✅ Input file uploaded: {input_file.name}")
            
            with st.expander("👁️ Preview Bank-Amount Data"):
                try:
                    file_content = input_file.read()
                    bank_data = read_bank_amount_data(file_content, input_file.name)
                    
                    if bank_data:
                        st.write(f"**Found {len(bank_data)} valid entries:**")
                        
                        # Create preview dataframe
                        preview_df = pd.DataFrame(bank_data[:10])  # Show first 10
                        if 'Date' in preview_df.columns:
                            preview_df['Date'] = preview_df['Date'].astype(str)
                        st.dataframe(preview_df, use_container_width=True)
                        
                        if len(bank_data) > 10:
                            st.write(f"... and {len(bank_data) - 10} more entries")
                        
                        # Show bank summary
                        banks = list(set(entry['Bank'] for entry in bank_data))
                        st.write(f"**Banks to search**: {', '.join(sorted(banks))}")
                        
                        st.session_state.bank_data = bank_data
                    else:
                        st.warning("No valid bank-amount-date entries found")
                        st.session_state.bank_data = []
                        
                except Exception as e:
                    st.error(f"Error reading input file: {e}")
                    st.session_state.bank_data = []
    
    with col2:
        st.subheader("📁 Source Files (Bank Statements)")
        
        source_files = st.file_uploader(
            "Upload bank statement files",
            type=['xlsx', 'xls'],
            accept_multiple_files=True,
            help="Upload Excel files for each bank. Filenames should contain bank names"
        )
        
        if source_files:
            st.success(f"✅ {len(source_files)} source file(s) uploaded")
            
            with st.expander("📋 Bank File Mapping"):
                if hasattr(st.session_state, 'bank_data') and st.session_state.bank_data:
                    banks_needed = list(set(entry['Bank'] for entry in st.session_state.bank_data))
                    
                    st.write("**Bank → File Mapping:**")
                    for bank in sorted(banks_needed):
                        matching_files = [f.name for f in source_files if check_bank_match(f.name, bank)]
                        if matching_files:
                            st.write(f"✅ **{bank}**: {', '.join(matching_files)}")
                        else:
                            st.write(f"❌ **{bank}**: No matching files found")
                else:
                    for file in source_files:
                        st.write(f"• **{file.name}**")
    
    # Search execution
    if input_file and source_files and hasattr(st.session_state, 'bank_data') and st.session_state.bank_data:
        
        st.subheader("🚀 Execute Bank-Specific Search")
        
        # Show search statistics
        col1, col2, col3 = st.columns(3)
        
        banks_count = len(set(entry['Bank'] for entry in st.session_state.bank_data))
        amounts_count = len(st.session_state.bank_data)
        files_count = len(source_files)
        
        with col1:
            st.metric("Unique Banks", banks_count)
        with col2:
            st.metric("Total Amounts", amounts_count)
        with col3:
            st.metric("Source Files", files_count)
        
        if st.button("🔍 Start Robust Bank Search", type="primary", use_container_width=True):
            
            st.session_state.search_log = ["=== ROBUST BANK SEARCH STARTED ==="]
            
            with st.spinner("Searching bank-specific amounts with robust processing..."):
                matches_dict, search_summary = process_bank_specific_search(
                    st.session_state.bank_data,
                    source_files,
                    tolerance
                )
            
            # Display results
            if matches_dict:
                total_matches = sum(len(matches) for matches in matches_dict.values())
                found_amounts = sum(1 for summary in search_summary.values() if summary['Status'] == 'Found')
                
                st.markdown(f"""
                <div class="success-box">
                🎉 <strong>Robust Search Complete!</strong><br>
                Found <strong>{total_matches}</strong> total matches across <strong>{len(matches_dict)}</strong> bank-file combinations<br>
                Successfully found <strong>{found_amounts}</strong> out of <strong>{len(st.session_state.bank_data)}</strong> searched amounts
                </div>
                """, unsafe_allow_html=True)
                
                # Quick summary
                st.subheader("📊 Quick Summary")
                summary_data = list(search_summary.values())
                if summary_data:
                    summary_df = pd.DataFrame(summary_data)
                    st.dataframe(summary_df, use_container_width=True)
                
                # Download
                st.subheader("💾 Download Excel Results")
                
                try:
                    with st.spinner("Creating Excel report with matches only..."):
                        excel_data = create_robust_excel_report(matches_dict, search_summary, st.session_state.bank_data)
                    
                    if excel_data:  # Only show download if file was created successfully
                        st.download_button(
                            label="📥 Download Results Only (No Summary Sheets)",
                            data=excel_data,
                            file_name=f"bank_search_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                        
                        st.markdown("""
                        <div class="success-box">
                        ✅ <strong>Excel Report Ready!</strong><br>
                        This version includes robust data cleaning and formatting to prevent Excel corruption issues.
                        The file contains only the result sheets with matches - no summary sheets.
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        st.error("Failed to create Excel file. Check the search log for details.")
                    
                except Exception as e:
                    st.error(f"Error creating download: {e}")
                    st.info("Check the search log for detailed error information.")
                
            else:
                st.warning("🔍 No matches found for any bank-amount combinations")
                st.info("💡 Since no matches were found, there are no result sheets to download.")
                
                with st.expander("💡 Troubleshooting"):
                    st.markdown("""
                    **No matches found? Check:**
                    - File names contain bank names (e.g., 'SNB.xlsx' for SNB bank)
                    - Amounts exist in the bank statement files
                    - Date filtering isn't too restrictive
                    - Tolerance value is appropriate
                    - Bank names in input match file naming convention
                    """)
    
    else:
        st.info("👆 Please upload both input file and source files to begin robust bank-specific search")

if __name__ == "__main__":
    main()
